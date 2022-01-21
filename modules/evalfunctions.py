# Criticality Evaluation and extraction of NOR/PAR-ranges from Excel Eval Sheets
# by Lukas Brändli, 2021

import pandas as pd

def importExcelData(filepath):
    #xlsm read
    from openpyxl import load_workbook
    from itertools import islice

    wb2 = load_workbook(filepath, read_only=True, data_only=True)
    # print(wb2.sheetnames)


    # dfs = {sheet_name: pd.DataFrame(wb2[sheet_name].values) 
    #         for sheet_name in wb2.sheetnames}

    dfs = {}

    for sheet_name in wb2.sheetnames:
        data = wb2[sheet_name].values
        cols = next(data)[1:]
        data = list(data)
        idx = [r[0] for r in data]
        data = (islice(r, 1, None) for r in data)
        df = pd.DataFrame(data, index=idx, columns=cols)

        dfs[sheet_name] = df

    wb2.close()


    return dfs



#Clean LABdata

def cleanLABdata(df):
    #Rename Columns
    df = df.rename(
        {
        'Spalte2':'Parameter',
        'ALL!_FilterDatabase':'Unit',
        'Experiment':'ParamType'
        },

        axis=1)

    #Delete unwanted Rows by Index
    df = df.drop(index=range(7))

    #Delete unwanted Columns by Name
    df = df.loc[:, ((df.columns.str.endswith('-00') == False)
               &(df.columns.str.contains('Unn') == False))]

    # df.drop(columns=['Param_Type'])
    df.Param_Type = df.Param_Type.fillna('-')

    df['Param'] = df['Parameter']

    return df


def cleanLABdataLIN(df):
    #Rename Columns
    
    df['Parameter'] = df.index
    df = df.reset_index(drop=True)
    
    
    df = df.rename(columns= {'ALL!_FilterDatabase':'Unit', 'Experiment':'ParamType'})


    df['Param'] = df['Parameter']

    return df



#Get numeric LAB data
def getNumericLABdata(df):
    
  
    df_long = pd.melt(df, id_vars=['Parameter', 'Param', 'Unit' , 'ParamType'], var_name='Experiment')
    df_long['value'] = pd.to_numeric(df_long['value'],errors='coerce') 
    df_long['value'] = df_long['value'].round(2)


    data_tidy = df_long.pivot_table(
       
        index=['Parameter','Param', 'Unit', 'ParamType'],

        #Columns = to be moved to new columns
        columns='Experiment',
        values='value'
    )

    

    data = data_tidy

    data.reset_index(inplace=True) 
    data_LAB = data.T
    data_LAB.columns = data_LAB.iloc[0,:]
    data_LAB['Experiment']=data_LAB.index
    data_LAB=data_LAB.iloc[1:,:]

    data_LAB = data_LAB.reset_index(drop=True)
   
    cols = data_LAB.columns.tolist()
    cols = cols[-1:] + cols[:-1]

    data_LAB = data_LAB[cols]
    data_LAB.columns.name = ''
    data_LAB.index.name = ''


    #Skip Header rows
    dataLABnumeric = data_LAB.iloc[4:,:]
   
    
    return dataLABnumeric



#Expand Process Parameters and Step Numbers
def getPARAMdata(df, meta=False):
    
    df=df.copy()
    
    df=df.loc[df['Parameter'].str.contains(':')]
    # df

    #Split Step from Parameter (separator, 1,expand=True) & EXPAND!
    splParam = df.Parameter.str.split(':', 1, expand=True)
    splParam

    #Reset & Drop Index
    splParam.reset_index(drop=True)
    splParam.shape

    # Extract and format Step Number
    splStep = splParam[0].str.split(' ', 1, expand=True)
    splStep[1] = pd.to_numeric(splStep[1], errors='coerce')
    
    #OK?
    splStep = splStep[splStep[1].isna() == False]

    #Insert leading 0 in Step Number
    splStep[1] = splStep[1].astype(int).apply(lambda x: '{0:0>2}'.format(x))
    # Combine in one Column
    splStep['Step']=splStep[0]+' '+splStep[1]

    df['Step'] = splStep['Step']
    df['Param'] = splParam[1]
    # df = df.drop(columns='Parameter')

    df = df.sort_values(by=['Step'],ascending=True)

    

    df['ParamType'] = df['ParamType'].fillna('-')

   

    df_long = pd.melt(df, id_vars=['Step', 'Parameter','Param', 'ParamType', 'Unit'], var_name=['Experiment'])
 
    #Convert values to numeric type for eval
    df_long['value'] = pd.to_numeric(df_long['value'], errors='coerce') 
    df_long['Experiment'].str.strip()
    
    if meta == True:
        df_long = df_long.drop(columns=['value','Experiment'])
        df_long = df_long.drop_duplicates('Parameter')
    
    return df_long

def getExtremes(PARAMdata, edge = 0.02):
    
    #Extract upper and under Extremes of process parameters
    
    PARAMedge = pd.DataFrame(PARAMdata.groupby(['Step','Param', 'ParamType','Unit']).quantile([edge,1-edge]).value)
    PARAMedge = PARAMedge.reset_index()
    PARAMedge = PARAMedge.rename({'level_4':'quantile', 'value':'edge'},axis=1)


    PARAMeval = PARAMdata.merge(PARAMedge, on=['Step','Param', 'ParamType', 'Unit'])
    upperExtremes = PARAMeval[PARAMeval['quantile'] > 0.5]
    lowerExtremes = PARAMeval[PARAMeval['quantile'] <= 0.5]

    lowerExtremes = lowerExtremes[lowerExtremes['value'] < lowerExtremes['edge']]
    upperExtremes = upperExtremes[upperExtremes['value'] >= upperExtremes['edge']]
    
    return lowerExtremes, upperExtremes


#get Experiment Descriptions (header)
def getExperimentDescription(dfs):   
    header = dfs['DATAProcess'].iloc[:7,:].copy().T
    header = header.reset_index()
    header = header.rename(columns={'index':'Experiment','Process Parameters':'Date'})
    header = header.iloc[3:,:]
    header = header[header['Experiment'].str.contains('BU')]
    header = header.fillna('-')
    
    return header



def getCDSpivot(CDSdata):
    #Select columns in relevant order
    # from datetime import datetime, date

    # for i in CDSdata.index:
    #     if type(CDSdata.loc[i,'data added']) == datetime:
    #         CDSdata.loc[i,'data added'] = str(CDSdata.loc[i,'data added'].date())
            

    CDSdata = CDSdata[['Sample', 'Prefix', 'Suffix', 'Peak Name ', 'RRT', 'Ret.Time',  'Rel.Area ', 'Area',
           'Height', 'Amount', 'Inj. Time', 'Sequence', 'Sequence Type', 'data added']]
    CDSdata.columns = CDSdata.columns.str.strip()

    #Rename 
    CDSdata=CDSdata.rename(columns={'Prefix':'Experiment','Suffix':'Sample Type'})
    CDSdata['Experiment'] = CDSdata['Experiment'].str.strip()
    CDSdata['Sample Type'] = CDSdata['Sample Type'].str.strip()

    #get rid of garbage
    CDSdata=CDSdata[
        (CDSdata['Experiment'].str.contains('BU'))
        &(CDSdata['Peak Name'].str.contains('RRT') | CDSdata['Peak Name'].str.contains('-'))               
                   ]

    id_vars=['Experiment', 'Sample Type', 'Peak Name', 'RRT', 'Ret.Time', 'Sequence Type', 'data added']
    # id_vars=['Experiment', 'Sample Type', 'Peak Name', 'RRT', 'Ret.Time', 'Sequence Type']

    CDSdata = CDSdata.melt(id_vars, 
                           value_vars=['Rel.Area', 'Area', 'Height', 'Amount'])

    CDSdata['value'] = pd.to_numeric(CDSdata['value'],errors='coerce')
    
    #Clean up Sample Types
    CDSdata['Sample Type'] = CDSdata['Sample Type'].str.strip()

    #CDSdata.loc[(CDSdata['Sample Type'] == 'cryst NM')]

    #Write data to PIVOT-table for structured output
    CDSpivot = CDSdata.pivot_table(
        #Index = columnns not to be moved
    #   index=['Sample','Experiment', 'Sample_Type', 'Peak Name '],
        index = id_vars,

        #Columns = to be moved to new columns
        columns='variable',
        values='value'
    )

    return CDSpivot

def getConsistency(CDSpivot, tolerance=0.03):
    completeData = CDSpivot.groupby(['Experiment', 'Sample Type']).sum()[['Rel.Area']]
    redundantData = completeData[(completeData['Rel.Area'] >= (100.00 + tolerance))]
    INcompleteData = completeData[(completeData['Rel.Area'] < (100.00 - tolerance))]
    completeData = completeData[(completeData['Rel.Area'] >= (100.00 - tolerance))
                 &(completeData['Rel.Area'] <= (100.00 + tolerance))
                ]

    consistencyCount = ['incomplete: ' + str(INcompleteData['Rel.Area'].count()),
    'redundant: ' + str(redundantData['Rel.Area'].count()),
    'complete: ' + str(completeData['Rel.Area'].count())]

    return consistencyCount, INcompleteData , redundantData, completeData 


#Get Specification
def getSPEC(dfs):
    SPEC = dfs['SPEC']
    SPEC = SPEC.iloc[1:5,:].T.reset_index(drop=True)
    SPEC = SPEC.rename(columns={'Compound':'Peak Name'})

    #remove entries without defined SPEC-limit
    SPEC['LSL'] = pd.to_numeric(SPEC['LSL'], errors='coerce')
    SPEC['USL'] = pd.to_numeric(SPEC['USL'], errors='coerce')
    SPEC = SPEC.dropna(axis=0)

    return SPEC



def getCRIT(CDSpivot, selSPEC, header, mode='compact', sample='cryst', allSamples=False):  
    #Select Sample
    
    CDSsel = CDSpivot[CDSpivot.index.get_level_values('Sample Type').isin([sample])]
    if (allSamples): CDSsel = CDSpivot

    CDSeval = pd.merge(CDSsel[['Rel.Area']].reset_index(), selSPEC, on='Peak Name')

    CDSeval[['Range']]=CDSeval.USL - CDSeval.LSL
    CDSeval[['CRIT']]=0.0

    #Criticality Calculation
    for x  in CDSeval.index:
        if CDSeval.loc[x, 'SELECT Critical Limit'] == 'USL':
               CDSeval.loc[x, 'CRIT'] = abs(1-(CDSeval.loc[x, 'USL'] - CDSeval.loc[x, 'Rel.Area'])/CDSeval.loc[x, 'Range'])

        if CDSeval.loc[x, 'SELECT Critical Limit'] == 'LSL':
               CDSeval.loc[x, 'CRIT'] = abs((CDSeval.loc[x, 'USL'] - CDSeval.loc[x, 'Rel.Area'])/CDSeval.loc[x, 'Range']) 

    #Merge with Experiment Metadata
    CDScrit = pd.merge(header, CDSeval, on='Experiment')

    #Write as PIVOT table
    CDScrit=CDScrit.pivot_table(
    index = ['Experiment', 'Sample Type', 'Aim'] ,

        #Columns = to be moved to new columns
        columns='Peak Name',
        values='CRIT'
    )

    
    CRITexp = list(CDScrit[CDScrit > 1].dropna(axis=1,how='all').index.get_level_values('Experiment'))
    
    #Drop Columns according to selected mode
    if (mode == 'full'):
        #CRIT > 0 only
        CDScrit = CDScrit[CDScrit > 0.0].dropna(axis=1,how='all')

    if (mode == 'compact'):
        #CRIT only
        CDScrit = CDScrit[CDScrit > 1].dropna(axis=1,how='all')
        CDScrit = CDScrit.dropna(how='all', axis=0)
        
   
    #TO BE DEFINED:
    if (mode == 'crit'):
        #CRIT only
        crit = CDScrit[CDScrit > 1]
        CRITimp = list(crit.dropna(axis=1,how='all').columns)
        CDScrit = CDScrit[CDScrit.columns[CDScrit.columns.isin(CRITimp)]].dropna(axis=1,how='all')
        
        #CDScrit = CDScrit.dropna(how='all', axis=0)
            
        
        
        
#     if (mode == 'uncrit'):
#         CDScrit = CDScrit[CDScrit <= 1].dropna(axis=1,how='all')
# #         CDScrit = CDScrit[~CDScrit.index.get_level_values('Experiment').isin(CRITexp)]
#         CDScrit = CDScrit[CDScrit.index.get_level_values('Experiment').isin(CDScrit.index.get_level_values('Experiment')[~CDScrit.index.get_level_values('Experiment').isin(CRITexp)])]

    
    #Filter CRIT experiments only
#     CDScrit = CDScrit.dropna(how='all', axis=0)

    
    #Maximum Criticality = Severity
    severity = CDScrit.max(axis=1)
    CDScrit['Severity'] = severity
    
    if (mode == 'uncrit'):
        #Remove Impurities with CRIT >1
        CDScrit = CDScrit[CDScrit['Severity'] <= 1].dropna(axis=1,how='all')
        #Remove entries with CRIT < 1 from Experiments Severity (CRITmax) > 1
        keep = CDScrit.index.get_level_values('Experiment')[~CDScrit.index.get_level_values('Experiment').isin(CRITexp)]
        CDScrit[CDScrit.index.get_level_values('Experiment').isin(keep)]
    
    CDScrit['Limit'] = 1

    return CDSeval, CDScrit


def predCRIT(data, selSPEC, header, mode='compact'):  
    
    # removed (, sample='cryst', allSamples=False)

    CDSeval = pd.merge(data, selSPEC, on='Peak Name')

    CDSeval[['Range']]=CDSeval.USL - CDSeval.LSL
    CDSeval[['CRIT']]=0.0

    #Criticality Calculation
    for x  in CDSeval.index:
        if CDSeval.loc[x, 'SELECT Critical Limit'] == 'USL':
               CDSeval.loc[x, 'CRIT'] = abs(1-(CDSeval.loc[x, 'USL'] - CDSeval.loc[x, 'pred'])/CDSeval.loc[x, 'Range'])

        if CDSeval.loc[x, 'SELECT Critical Limit'] == 'LSL':
               CDSeval.loc[x, 'CRIT'] = abs((CDSeval.loc[x, 'USL'] - CDSeval.loc[x, 'pred'])/CDSeval.loc[x, 'Range']) 

    #Merge with Experiment Metadata
    CDScrit = pd.merge(header, CDSeval, on='Experiment')

    #Write as PIVOT table
    CDScrit=CDScrit.pivot_table(
    index = ['Experiment', 'Sample Type', 'Aim'] ,

        #Columns = to be moved to new columns
        columns='Peak Name',
        values='CRIT'
    )

    
    CRITexp = list(CDScrit[CDScrit > 1].dropna(axis=1,how='all').index.get_level_values('Experiment'))
    
    #Drop Columns according to selected mode
    if (mode == 'full'):
        #CRIT > 0 only
        CDScrit = CDScrit[CDScrit > 0.0].dropna(axis=1,how='all')

    if (mode == 'compact'):
        #CRIT only
        CDScrit = CDScrit[CDScrit > 1].dropna(axis=1,how='all')
        CDScrit = CDScrit.dropna(how='all', axis=0)
        
   
    #TO BE DEFINED:
    if (mode == 'crit'):
        #CRIT only
        crit = CDScrit[CDScrit > 1]
        CRITimp = list(crit.dropna(axis=1,how='all').columns)
        CDScrit = CDScrit[CDScrit.columns[CDScrit.columns.isin(CRITimp)]].dropna(axis=1,how='all')
        
        #CDScrit = CDScrit.dropna(how='all', axis=0)
            
        
        
        
#     if (mode == 'uncrit'):
#         CDScrit = CDScrit[CDScrit <= 1].dropna(axis=1,how='all')
# #         CDScrit = CDScrit[~CDScrit.index.get_level_values('Experiment').isin(CRITexp)]
#         CDScrit = CDScrit[CDScrit.index.get_level_values('Experiment').isin(CDScrit.index.get_level_values('Experiment')[~CDScrit.index.get_level_values('Experiment').isin(CRITexp)])]

    
    #Filter CRIT experiments only
#     CDScrit = CDScrit.dropna(how='all', axis=0)

    
    #Maximum Criticality = Severity
    severity = CDScrit.max(axis=1)
    CDScrit['Severity'] = severity
    
    if (mode == 'uncrit'):
        #Remove Impurities with CRIT >1
        CDScrit = CDScrit[CDScrit['Severity'] <= 1].dropna(axis=1,how='all')
        #Remove entries with CRIT < 1 from Experiments Severity (CRITmax) > 1
        keep = CDScrit.index.get_level_values('Experiment')[~CDScrit.index.get_level_values('Experiment').isin(CRITexp)]
        CDScrit[CDScrit.index.get_level_values('Experiment').isin(keep)]
    
    CDScrit['Limit'] = 1

    return CDSeval, CDScrit


def pad_dict_list(dict_list, padel=''):
    # Expend lists in dict to same length for combination in Dataframe
    # padel: fill text (usually '')
    
    lmax = 0
    for lname in dict_list.keys():
        lmax = max(lmax, len(dict_list[lname]))
    for lname in dict_list.keys():
        ll = len(dict_list[lname])
        if  ll < lmax:
            dict_list[lname] += [padel] * (lmax - ll)
    return dict_list



def formatHTMLtable(data, filename):

    data.to_html("./html/dashboard/rawHTML/"+ filename, classes = 'mystyle', index=False, index_names=False, sparsify=True)
    source = "./html/dashboard/rawHTML/" + filename
    target = "./html/dashboard/" + filename

    f = open(source,'r')
    HTMLtable = f.read()
    f.close()

    tableCSShead = '<html><head><title>HTML Pandas Dataframe with CSS</title></head><link rel="stylesheet" type="text/css" href="df_style.css"/><body>'
    tableCSSfoot = '</body></html>'

    formattedTable = tableCSShead + HTMLtable + tableCSSfoot

    f = open(target,'w')
    f.write(formattedTable)
    f.close()


"""  #Under Construction...
def getIPCcrit(CDSpivot, spec=SPEC, USL=smUSL, LSL=smLSL, header, mode='compact', sample='cryst', allSamples=False):  
    #Select Sample
    
    ipcSPEC = spec.drop(columns=['USL', 'LSL'])
    ipcSPEC = pd.merge(ipcSPEC, USL[['USL']], on='Peak Name', how = 'inner')
    
    CDSsel = CDSpivot[CDSpivot.index.get_level_values('Sample Type').isin([sample])]
    if (allSamples): CDSsel = CDSpivot

    CDSeval = pd.merge(CDSsel[['Rel.Area']].reset_index(), SPEC, on='Peak Name')

    CDSeval[['Range']]=CDSeval.USL - CDSeval.LSL
    CDSeval[['CRIT']]=0.0

    #Criticality Calculation
    for x  in CDSeval.index:
        if CDSeval.loc[x, 'SELECT Critical Limit'] == 'USL':
               CDSeval.loc[x, 'CRIT'] = abs(1-(CDSeval.loc[x, 'USL'] - CDSeval.loc[x, 'Rel.Area'])/CDSeval.loc[x, 'Range'])

        if CDSeval.loc[x, 'SELECT Critical Limit'] == 'LSL':
               CDSeval.loc[x, 'CRIT'] = abs((CDSeval.loc[x, 'USL'] - CDSeval.loc[x, 'Rel.Area'])/CDSeval.loc[x, 'Range']) 

    #Merge with Experiment Metadata
    CDScrit = pd.merge(header, CDSeval, on='Experiment')

    #Write as PIVOT table
    CDScrit=CDScrit.pivot_table(
    index = ['Experiment', 'Sample Type', 'Aim'] ,

        #Columns = to be moved to new columns
        columns='Peak Name',
        values='CRIT'
    )

    
    CRITexp = list(CDScrit[CDScrit > 1].dropna(axis=1,how='all').index.get_level_values('Experiment'))
    
    #Drop Columns according to selected mode
    if (mode == 'full'):
        #CRIT > 0 only
        CDScrit = CDScrit[CDScrit > 0.0].dropna(axis=1,how='all')

    if (mode == 'compact'):
        #CRIT only
        CDScrit = CDScrit[CDScrit > 1].dropna(axis=1,how='all')
        CDScrit = CDScrit.dropna(how='all', axis=0)
        
   
    #TO BE DEFINED:
    if (mode == 'crit'):
        #CRIT only
        crit = CDScrit[CDScrit > 1]
        CRITimp = list(crit.dropna(axis=1,how='all').columns)
        CDScrit = CDScrit[CDScrit.columns[CDScrit.columns.isin(CRITimp)]].dropna(axis=1,how='all')
        
        #CDScrit = CDScrit.dropna(how='all', axis=0)
            
        
        
        
#     if (mode == 'uncrit'):
#         CDScrit = CDScrit[CDScrit <= 1].dropna(axis=1,how='all')
# #         CDScrit = CDScrit[~CDScrit.index.get_level_values('Experiment').isin(CRITexp)]
#         CDScrit = CDScrit[CDScrit.index.get_level_values('Experiment').isin(CDScrit.index.get_level_values('Experiment')[~CDScrit.index.get_level_values('Experiment').isin(CRITexp)])]

    
    #Filter CRIT experiments only
#     CDScrit = CDScrit.dropna(how='all', axis=0)

    
    #Maximum Criticality = Severity
    severity = CDScrit.max(axis=1)
    CDScrit['Severity'] = severity
    
    if (mode == 'uncrit'):
        #Remove Impurities with CRIT >1
        CDScrit = CDScrit[CDScrit['Severity'] <= 1].dropna(axis=1,how='all')
        #Remove entries with CRIT < 1 from Experiments Severity (CRITmax) > 1
        keep = CDScrit.index.get_level_values('Experiment')[~CDScrit.index.get_level_values('Experiment').isin(CRITexp)]
        CDScrit[CDScrit.index.get_level_values('Experiment').isin(keep)]
    
    CDScrit['Limit'] = 1

    return CDSeval, CDScrit
"""